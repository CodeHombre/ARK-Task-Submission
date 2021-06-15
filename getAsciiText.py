import openpyxl
import sys



wb = openpyxl.load_workbook("pixelVals.xlsx")
ws = wb.active
rows = ws.max_row
cols = ws.max_column
for i in range(1,11):
    for j in range(1,cols+1):
        asciiVal = ws.cell(row=i,column=j).value
        ws.cell(row=i,column=j).value = ""
        print(chr(asciiVal),end='')
        if(chr(asciiVal)==':'):
            wb.save("decodedVals.xlsx")
            print("\n\n")
            print(i,j)
            sys.exit()
        
 
